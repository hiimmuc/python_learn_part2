import funtion_
import classes
import openpyxl as xl
from pathlib import Path
from openpyxl.chart import BarChart, PieChart, Reference

# exercise 1
# fizz buzz game


# print(fizz_buzz(int(input("Enter number: "))))

# exercise 2
# guessing game

# secret_word = "Nam"
# guess = ""
# guess_count = 0
# guess_limit = 3
# out_of_guess = False
# while guess != secret_word and not out_of_guess:
#     guess_count += 1
#     if guess_count <= guess_limit:
#         guess = input("enter word: ")
#         print("wrong!")
#     else:
#         out_of_guess = True
# if out_of_guess:
#     print("You lose the game")
# else:
#     print("you win")

# exercise 3
# car game

# answer = "nothing"
# started = False
# while answer:
#     answer = input(">>> ").lower()
#     if answer == "start":
#         if started:
#             print("car is already started!")
#         else:
#             started = True
#             print("Car is starting ... ")
#     elif answer == "stop":
#         if not started:
#             print("Car is already stopped!")
#         else:
#             started = False
#             print("Car is stopping")
#     elif answer == "quit":
#         answer = False
#     elif answer == "help":
#         print("type:\nstart - to start the car"
#               "\nstop - to stop the car"
#               "\nquit - to quit the game")
#     else:
#         print("invalid input")

number = {
    "1": "One ",
    "2": "Two ",
    "3": "Three ",
    "4": "Four ",
    "5": "Five ",
    "6": "Six ",
    "7": "Seven ",
    "8": "Eight ",
    "9": "Nine ",
    "0": "Zero "
}

# emoji converter

# sentence = input(">> ")
# print(funtion_.emoji_converter(sentence))

coordinate_ = (1, 2, 3)
vector1 = classes.Vector(1, 2, 3)


# vector1.print_value()

# path = Path()
# for file in path.glob('*.py'):
#     print(file)

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        print(type(corrected_price))

    value = Reference(sheet, min_row=2,
                      max_row=sheet.max_row,
                      min_col=4,
                      max_col=4)
    chart = PieChart()
    chart.add_data(value)
    sheet.add_chart(chart, 'a2')
    wb.save(filename)
    wb.close()


lib_1 = {'Harry Potter', 'Hunger Game', 'Lord of the Rings'}
lib_2 = {'Harry Potter', 'Romeo and Juliet'}

print(lib_1.union(lib_2))
